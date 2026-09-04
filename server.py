# serve.py
import os
import io
from fastapi.middleware.cors import CORSMiddleware
import json
import re
import datetime as dt
from typing import Optional, List, Dict
from fastapi import FastAPI, HTTPException, Query, BackgroundTasks, Depends, Header, Request, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from supabase import create_client, Client
from datetime import datetime, timedelta
from collections import defaultdict
import time
import csv
from models import Import, Notice, AuthCode, TendernedRawCached, TendernedRawCPVCached, TendernedRaw, SROIResult, NoticeSROIResult, CRMCompany, CRMFollowup, TenderNotice, TenderCompany, TenderNoticeCPV, TenderNoticeLot, SavedSearch, SavedSearchAlert

from sqlalchemy.orm import Session
from database import get_db, SessionLocal
import random

from sqlalchemy import desc, func, and_, or_

from dotenv import load_dotenv
from sroi_scanner import analyze_notice_sroi




# importeer run_import en helpers uit je scraper-bestand:
from final_tenderned import run_import  # pas de bestandsnaam aan
from sroi_scanner import analyze_import_sroi  # nieuwe SROI analyzer

# Load .env file
load_dotenv()

app = FastAPI(title="TenderNed Import Backend")

# In-memory store voor SROI analysis progress tracking
sroi_analysis_status: Dict[str, Dict] = {}



# 1. CORS - Only allow your frontend domain
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://vercel.com/ferfs-projects/ithaka/5vz8wR52b8fTVTXtxzxBWYT7DMa6",  # Replace with your actual domain
        "http://localhost:3000",              # For local dev
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 2. Rate Limiter
class RateLimiter:
    def __init__(self, max_requests=100, window_seconds=60):
        self.max_requests = max_requests
        self.window_seconds = window_seconds
        self.requests = defaultdict(list)
    
    def is_allowed(self, ip: str) -> bool:
        now = time.time()
        self.requests[ip] = [t for t in self.requests[ip] 
                            if now - t < self.window_seconds]
        
        if len(self.requests[ip]) >= self.max_requests:
            return False
        
        self.requests[ip].append(now)
        return True

rate_limiter = RateLimiter(max_requests=100, window_seconds=60)


@app.on_event("startup")
def fail_interrupted_imports() -> None:
    """Make jobs interrupted by a process restart visible instead of stuck."""
    db = SessionLocal()
    try:
        interrupted = db.query(Import).filter(Import.status == "running").update(
            {
                Import.status: "failed",
                Import.error_message: "Import interrupted because the backend restarted.",
                Import.finished_at: dt.datetime.now(dt.timezone.utc),
            },
            synchronize_session=False,
        )
        db.commit()
        if interrupted:
            print(f"⚠️ Marked {interrupted} interrupted import(s) as failed after startup")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    # Get real IP from Cloudflare
    client_ip = request.headers.get("CF-Connecting-IP") or request.client.host
    
    if not rate_limiter.is_allowed(client_ip):
        raise HTTPException(status_code=429, detail="Too many requests")
    
    response = await call_next(request)
    return response



# --------- Pydantic modellen ---------
class ImportRequest(BaseModel):
    date_from: str
    date_to: str
    publicatie_type: Optional[str] = "AGO"
    cpv_codes: Optional[List[str]] = None
    max_pages: Optional[int] = None
    region: Optional[str] = None  # ✅ Add this



class ImportResponse(BaseModel):
    import_id: str
    name: str
    total_records: int
    created_at: str
    # Fetch counters so callers can tell a complete import from a truncated one.
    # complete=False means some publications were listed but never stored.
    listed: int = 0
    fetched: int = 0
    http_failed: int = 0
    parse_failed: int = 0
    complete: bool = True


class ImportJobResponse(BaseModel):
    """Acknowledgement returned as soon as a queued import has been persisted."""
    import_id: str
    name: str
    status: str
    created_at: Optional[str] = None


class SingleNoticeSROIRequest(BaseModel):
    """Optional body for POST /imports/{id}/notices/{nid}/sroi-analyze."""
    target: str = "winner"


class SROIAnalysisStatus(BaseModel):
    import_id: str
    status: str  # "pending", "running", "completed", "failed"
    progress: int  # 0-100
    current: int
    total: int
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    results_summary: Optional[Dict] = None
    error: Optional[str] = None


# --------- Helpers ---------
def generate_import_name() -> str:
    now = dt.datetime.utcnow().replace(microsecond=0)
    return "import_" + now.isoformat().replace(":", "-")


# ----------------------
# Simple header-based auth
# Expects header `X-User-Code: 123456789012` (12 digits)
# ----------------------

def validate_user_code(
    x_user_code: str = Header(..., alias="X-User-Code"),
    db: Session = Depends(get_db)
) -> str:
    """Validate that the provided header is a 12-digit numeric code.
    Returns the code when valid, otherwise raises HTTPException(401).
    """
    if not x_user_code:
        raise HTTPException(status_code=401, detail="X-User-Code header missing")
    
    code = str(x_user_code).strip()
    if len(code) != 12 or not code.isdigit():
        raise HTTPException(status_code=401, detail="Invalid user code. Must be 12 digits.")
    
    # Server-side validation: ensure the code exists in the whitelist table
    try:
        auth_code = db.query(AuthCode).filter(AuthCode.code == code).first()
        if not auth_code:
            raise HTTPException(
                status_code=403, 
                detail="Unknown user code; please register your code or contact admin."
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to validate user code: {e}")
    
    return code





def ensure_sroi_table():
    """
    Maak SROI results tabel als die niet bestaat.
    Supabase SQL migrations zijn beter, maar dit werkt voor snelle setup.
    """
    # Note: Je moet deze tabel handmatig aanmaken in Supabase dashboard:
    """
    CREATE TABLE IF NOT EXISTS sroi_results (
        id BIGSERIAL PRIMARY KEY,
        import_id TEXT NOT NULL REFERENCES imports(id) ON DELETE CASCADE,
        notice_id TEXT,
        publicatie_id TEXT,
        winner_name TEXT,
        analyzed_url TEXT,
        url_source TEXT,
        sroi_compliant BOOLEAN DEFAULT FALSE,
        confidence TEXT,
        score INTEGER DEFAULT 0,
        evidence JSONB,
        summary TEXT,
        pages_checked INTEGER DEFAULT 0,
        error TEXT,
        created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW()
    );
    
    CREATE INDEX IF NOT EXISTS idx_sroi_import_id ON sroi_results(import_id);
    CREATE INDEX IF NOT EXISTS idx_sroi_compliant ON sroi_results(sroi_compliant);
    """
    pass


# -----------------------------------------------
# Region / Province mapping helpers
# -----------------------------------------------
# A small mapping of Dutch provinces to example cities. This
# is not exhaustive but covers common cities seen in the dataset.
PROVINCE_TO_CITIES = {
    "Noord-Holland": [
        "Amsterdam", "Haarlem", "Alkmaar", "Zaanstad", "Hilversum", "Purmerend", "Hoorn", "Beverwijk", "Diemen", "Velsen",
        "Amstelveen", "Haarlemmermeer", "Heerhugowaard", "Langedijk", "Medemblik", "Schagen", "Enkhuizen", "Edam-Volendam",
        "Texel", "Bussum", "Huizen", "Blaricum", "Naarden", "Weesp", "Muiden", "Heemstede", "Bloemendaal", "Zandvoort",
        "Castricum", "Uitgeest", "Heiloo", "Bergen", "Ouder-Amstel", "Uithoorn", "Aalsmeer", "Landsmeer", "Waterland",
        "Oostzaan", "Wormerland", "Beemster", "Zeevang", "Koggenland", "Stede Broec", "Drechterland", "Opmeer", "Hollands Kroon",
        "Den Helder", "Heemskerk", "Uitgeest"
    ],
    "Zuid-Holland": [
        "Rotterdam", "Den Haag", "Leiden", "Delft", "Zoetermeer", "Dordrecht", "Gouda", "Schiedam", "Spijkenisse", "Vlaardingen",
        "Capelle aan den IJssel", "Alphen aan den Rijn", "Rijswijk", "Leidschendam-Voorburg", "Westland", "Katwijk", "Maassluis",
        "Krimpen aan den IJssel", "Pijnacker-Nootdorp", "Papendrecht", "Ridderkerk", "Hellevoetsluis", "Nissewaard", "Waddinxveen",
        "Nieuwkoop", "Voorschoten", "Wassenaar", "Zoeterwoude", "Lisse", "Noordwijk", "Teylingen", "Hillegom", "Oegstgeest",
        "Bodegraven-Reeuwijk", "Zuidplas", "Krimpenerwaard", "Alblasserdam", "Sliedrecht", "Zwijndrecht", "Hendrik-Ido-Ambacht",
        "Barendrecht", "Albrandswaard", "Lansingerland", "Westvoorne", "Brielle", "Goeree-Overflakkee", "Molenlanden",
        "Hoeksche Waard", "Gorinchem", "Hardinxveld-Giessendam", "Giessenlanden"
    ],
    "Utrecht": [
        "Utrecht", "Amersfoort", "Nieuwegein", "Veenendaal", "Zeist", "Houten", "Soest",
        "IJsselstein", "Woerden", "Wijk bij Duurstede", "De Bilt", "Bunnik", "Stichtse Vecht",
        "Vijfheerenlanden", "Montfoort", "Oudewater", "Lopik", "Renswoude", "Rhenen", "Woudenberg",
        "Leusden", "Bunschoten", "Baarn", "Eemnes", "De Ronde Venen", "Utrechtse Heuvelrug"
    ],
    "Noord-Brabant": [
        "Eindhoven", "Breda", "Tilburg", "s-Hertogenbosch", "Helmond", "Roosendaal", "Oss", "Bergen op Zoom", "Waalwijk",
        "Veldhoven", "Oosterhout", "Etten-Leur", "Uden", "Best", "Veghel", "Valkenswaard", "Boxmeer", "Geldrop-Mierlo",
        "Cuijk", "Heusden", "Boxtel", "Bernheze", "Eersel", "Oirschot", "Bladel", "Reusel-De Mierden", "Hilvarenbeek",
        "Goirle", "Oisterwijk", "Baarle-Nassau", "Alphen-Chaam", "Gilze en Rijen", "Dongen", "Loon op Zand", "Heeze-Leende",
        "Cranendonck", "Nuenen", "Son en Breugel", "Deurne", "Asten", "Someren", "Laarbeek", "Gemert-Bakel", "Sint-Michielsgestel",
        "Schijndel", "Meierijstad", "Landerd", "Boekel", "Mill en Sint Hubert", "Sint Anthonis", "Grave", "Wijchen", "Drimmelen",
        "Geertruidenberg", "Altena", "Werkendam", "Woudrichem", "Steenbergen", "Moerdijk", "Halderberge", "Zundert", "Rucphen",
        "Oosterhout", "Vught", "Oisterwijk"
    ],
    "Gelderland": [
        "Arnhem", "Nijmegen", "Apeldoorn", "Ede", "Doetinchem", "Zutphen",
        "Harderwijk", "Tiel", "Wageningen", "Zevenaar", "Duiven", "Barneveld", "Winterswijk", "Aalten", "Berkelland",
        "Bronckhorst", "Montferland", "Oude IJsselstreek", "Lochem", "Voorst", "Epe", "Hattem", "Heerde", "Oldebroek",
        "Elburg", "Ermelo", "Nunspeet", "Putten", "Nijkerk", "Scherpenzeel", "Renkum", "Rheden", "Rozendaal", "Lingewaard",
        "Overbetuwe", "West Maas en Waal", "Druten", "Wijchen", "Beuningen", "Heumen", "Mook en Middelaar", "Groesbeek",
        "Berg en Dal", "Culemborg", "Buren", "Neder-Betuwe", "West Betuwe", "Lingewaal", "Neerijnen", "Zaltbommel",
        "Maasdriel", "Tiel", "Brummen", "Westervoort", "Zevenaar", "Doesburg", "Oost Gelre", "Montferland", "Oude IJsselstreek"
    ],
    "Overijssel": [
        "Enschede", "Zwolle", "Deventer", "Almelo", "Hengelo", "Oldenzaal",
        "Kampen", "Hardenberg", "Dalfsen", "Raalte", "Hellendoorn", "Wierden", "Rijssen-Holten", "Borne", "Losser",
        "Haaksbergen", "Tubbergen", "Dinkelland", "Twenterand", "Olst-Wijhe", "Staphorst", "Zwartewaterland", "Steenwijkerland",
        "Ommen", "Olst-Wijhe"
    ],
    "Limburg": [
        "Maastricht", "Venlo", "Sittard-Geleen", "Heerlen", "Roermond", "Weert",
        "Venray", "Kerkrade", "Brunssum", "Landgraaf", "Voerendaal", "Simpelveld", "Vaals", "Gulpen-Wittem", "Valkenburg aan de Geul",
        "Eijsden-Margraten", "Maastricht", "Meerssen", "Beek", "Stein", "Meerssen", "Nederweert", "Leudal", "Echt-Susteren",
        "Maasgouw", "Roerdalen", "Peel en Maas", "Horst aan de Maas", "Venray", "Bergen", "Gennep", "Beesel", "Mook en Middelaar"
    ],
    "Friesland": [
        "Leeuwarden", "Heerenveen", "Drachten", "Sneek",
        "Smallingerland", "Súdwest-Fryslân", "Harlingen", "Franeker", "Dokkum", "Bolsward", "Lemsterland", "Skarsterlân",
        "Opsterland", "Ooststellingwerf", "Weststellingwerf", "Achtkarspelen", "Tytsjerksteradiel", "Kollumerland en Nieuwkruisland",
        "Dantumadiel", "Ferwerderadiel", "Menameradiel", "Het Bildt", "Franekeradeel", "Littenseradiel", "Wymbritseradiel",
        "Gaasterland-Sloten", "Lemsterland", "Nijefurd", "De Fryske Marren", "Waadhoeke", "Noardeast-Fryslân", "Ameland", "Schiermonnikoog", "Terschelling", "Vlieland"
    ],
    "Groningen": [
        "Groningen", "Delfzijl", "Hoogezand-Sappemeer",
        "Veendam", "Stadskanaal", "Winschoten", "Appingedam", "Pekela", "Oldambt", "Menterwolde", "Slochteren", "Haren",
        "Ten Boer", "Bedum", "Winsum", "De Marne", "Eemsmond", "Loppersum", "Zuidhorn", "Leek", "Marum", "Grootegast",
        "Vlagtwedde", "Bellingwedde", "Westerwolde", "Het Hoogeland", "Westerkwartier", "Midden-Groningen", "Eemsdelta"
    ],
    "Drenthe": [
        "Assen", "Emmen", "Hoogeveen",
        "Meppel", "Coevorden", "Borger-Odoorn", "Aa en Hunze", "Noordenveld", "Tynaarlo", "Midden-Drenthe", "Westerveld",
        "De Wolden"
    ],
    "Flevoland": [
        "Almere", "Lelystad", "Dronten",
        "Zeewolde", "Urk", "Noordoostpolder"
    ],
    "Zeeland": [
        "Middelburg", "Vlissingen", "Goes", "Terneuzen",
        "Hulst", "Sluis", "Borsele", "Kapelle", "Reimerswaal", "Veere", "Schouwen-Duiveland", "Tholen", "Noord-Beveland"
    ],
}

# Build reverse mapping city -> province for fast lookup (normalize to lower-case)
CITY_TO_PROVINCE = {}
for prov, cities in PROVINCE_TO_CITIES.items():
    for c in cities:
        CITY_TO_PROVINCE[c.lower()] = prov


def map_city_to_province(city: Optional[str]) -> Optional[str]:
    """
    Return province name for a given city string (case-insensitive).
    If city contains extra text (e.g. 'Amsterdam (NL)') we try to match substring.
    """
    if not city:
        return None
    city_norm = city.strip().lower()

    # Direct lookup
    if city_norm in CITY_TO_PROVINCE:
        return CITY_TO_PROVINCE[city_norm]

    # Try to match by startswith or contains for multi-word fields
    for known_city, prov in CITY_TO_PROVINCE.items():
        if known_city in city_norm:
            return prov

    return None


def map_postcode_to_province(postcode: Optional[str]) -> Optional[str]:
    """Dutch postcode ranges provide a safer fallback than matching a city name."""
    if not postcode:
        return None
    match = re.match(r"\s*(\d{4})", postcode)
    if not match:
        return None
    number = int(match.group(1))
    # Postal ranges overlap at municipal boundaries, so this is deliberately a
    # fallback only. A known municipality/city still wins in province_for_record.
    ranges = (
        (1000, 1299, "Noord-Holland"), (1300, 1379, "Flevoland"),
        (1380, 1499, "Noord-Holland"), (1500, 2199, "Zuid-Holland"),
        (2200, 2299, "Zuid-Holland"), (2300, 2399, "Zuid-Holland"),
        (2400, 2999, "Zuid-Holland"), (3000, 3799, "Zuid-Holland"),
        (3800, 3999, "Utrecht"), (4000, 4299, "Gelderland"),
        (4300, 4999, "Zeeland"), (5000, 5999, "Noord-Brabant"),
        (6000, 6599, "Limburg"), (6600, 7399, "Gelderland"),
        (7400, 7799, "Overijssel"), (7800, 7999, "Drenthe"),
        (8000, 8299, "Overijssel"), (8300, 8999, "Friesland"),
        (9000, 9999, "Groningen"),
    )
    for start, end, province in ranges:
        if start <= number <= end:
            return province
    return None


def province_for_record(record: dict) -> Optional[str]:
    return map_city_to_province(record.get("win_plaats")) or map_postcode_to_province(record.get("win_postcode"))


def _parse_iso_date(value):
    if not value:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    try:
        return dt.date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _company_key(name: Optional[str], kvk: Optional[str], postcode: Optional[str]) -> Optional[str]:
    if kvk:
        return "kvk:" + re.sub(r"\W", "", kvk).lower()
    if not name:
        return None
    return "name:" + re.sub(r"\W", "", name).lower() + ":" + re.sub(r"\W", "", postcode or "").lower()


def _upsert_tender_company(record: dict, prefix: str, db: Session) -> Optional[TenderCompany]:
    name = record.get(f"{prefix}_bedrijf_naam")
    kvk = record.get(f"{prefix}_kvk")
    key = _company_key(name, kvk, record.get(f"{prefix}_postcode"))
    if not key or not name:
        return None
    company = db.query(TenderCompany).filter(TenderCompany.canonical_key == key).first()
    if company is None and kvk:
        company = db.query(TenderCompany).filter(TenderCompany.kvk == kvk).first()
    province = map_city_to_province(record.get(f"{prefix}_plaats")) or map_postcode_to_province(record.get(f"{prefix}_postcode"))
    if company is None:
        company = TenderCompany(canonical_key=key, name=name)
        db.add(company)
    company.kvk = kvk or company.kvk
    company.name = name or company.name
    company.street = record.get(f"{prefix}_straat") or company.street
    company.postcode = record.get(f"{prefix}_postcode") or company.postcode
    company.city = record.get(f"{prefix}_plaats") or company.city
    company.province = province or company.province
    company.country = record.get(f"{prefix}_land") or company.country
    company.website = record.get(f"{prefix}_website") or company.website
    company.last_seen_at = dt.datetime.now(dt.timezone.utc)
    db.flush()
    return company


def persist_canonical_records(records: list[dict], db: Session) -> None:
    """Persist source payloads and normalized CPV/company/lot records atomically."""
    for record in records:
        notice_key = record.get("notice_id") or str(record.get("publicatieId") or "")
        source_xml = record.get("_source_xml")
        if not notice_key or not source_xml:
            continue
        notice = db.query(TenderNotice).filter(TenderNotice.notice_id == notice_key).first()
        if notice is None:
            notice = TenderNotice(notice_id=notice_key, source_xml=source_xml)
            db.add(notice)
        winner = _upsert_tender_company(record, "win", db)
        buyer = _upsert_tender_company(record, "buyer", db)
        parsed = {key: value for key, value in record.items() if not key.startswith("_")}
        notice.publicatie_id = record.get("publicatieId") or record.get("publicatie_id")
        notice.publicatie_datum = _parse_iso_date(record.get("publicatie_datum") or record.get("contract_issue_date"))
        notice.publicatie_code = record.get("publicatie_code")
        notice.type_publicatie = record.get("type_publicatie")
        notice.record_type = record.get("record_type")
        notice.is_cancelled = bool(record.get("is_cancelled"))
        notice.title = record.get("titel")
        notice.description = record.get("omschrijving")
        source_url = record.get("URL") or record.get("url")
        if isinstance(source_url, dict):
            source_url = source_url.get("href") or source_url.get("url")
        notice.source_url = source_url
        notice.listing_payload = json.loads(json.dumps(record.get("_listing_payload") or {}, default=str))
        notice.source_xml = source_xml
        notice.parsed_payload = json.loads(json.dumps(parsed, default=str))
        notice.winner_company = winner
        notice.buyer_company = buyer
        notice.updated_at = dt.datetime.now(dt.timezone.utc)
        db.flush()

        db.query(TenderNoticeCPV).filter(TenderNoticeCPV.notice_id == notice.id).delete()
        for index, code in enumerate(record.get("cpv_codes") or []):
            db.add(TenderNoticeCPV(notice_id=notice.id, code=str(code), is_main=index == 0))

        db.query(TenderNoticeLot).filter(TenderNoticeLot.notice_id == notice.id).delete()
        for lot in record.get("lots") or []:
            lot_id = lot.get("external_lot_id")
            if not lot_id:
                continue
            db.add(TenderNoticeLot(
                notice_id=notice.id,
                external_lot_id=str(lot_id),
                title=lot.get("title"),
                description=lot.get("description"),
                award_value=lot.get("award_value"),
                currency=lot.get("currency"),
                award_date=_parse_iso_date(lot.get("award_date")),
                contract_start=_parse_iso_date(lot.get("contract_start")),
                contract_end=_parse_iso_date(lot.get("contract_end")),
            ))
    db.commit()


# --------- Endpoints ---------

@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/imports")
def list_imports(
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Lijst eerdere imports (voor UI-overzicht).
    """
    imports = db.query(Import) \
        .filter(Import.owner_code == user_code) \
        .order_by(desc(Import.created_at)) \
        .limit(50) \
        .all()
    
    # Convert to dict format
    return [
        {
            "id": imp.id,
            "name": imp.name,
            "date_from": imp.date_from,
            "date_to": imp.date_to,
            "publicatie_type": imp.publicatie_type,
            "cpv_codes": imp.cpv_codes,
            "region": imp.region,
            "total_records": imp.total_records,
            "owner_code": imp.owner_code,
            "created_at": imp.created_at,
            # Lifecycle + completeness, so the UI can distinguish a finished
            # import from a crashed or truncated one.
            "status": imp.status,
            "error_message": imp.error_message,
            "started_at": imp.started_at,
            "finished_at": imp.finished_at,
            "listed": imp.listed,
            "fetched": imp.fetched,
            "http_failed": imp.http_failed,
            "parse_failed": imp.parse_failed,
            "complete": (imp.http_failed + imp.parse_failed) == 0,
        }
        for imp in imports
    ]


BATCH_SIZE = 1000

def map_import_to_cache_schema(record: dict, include_cpv: bool = False) -> dict:
    """
    Map een import record naar het cache schema.
    Inclusief historische velden voor consistentie.
    """
    # Extract URL - handle both dict and string formats
    url_value = record.get("URL") or record.get("url")
    if isinstance(url_value, dict):
        url_value = url_value.get("href") or url_value.get("url")
    
    # Get publication date from any available field
    pub_date = (
        record.get("contract_issue_date") or 
        record.get("publicatie_datum") or 
        record.get("Publicatiedatum")  # ← Add this!
    )
    
    mapped = {
        "notice_id": record.get("notice_id"),
        "publicatie_id": record.get("publicatieId") or record.get("publicatie_id"),
        "url": url_value,
        "titel": record.get("titel"),
        "omschrijving": record.get("omschrijving"),
        "win_bedrijf_naam": record.get("win_bedrijf_naam"),
        "win_kvk": record.get("win_kvk"),
        "win_straat": record.get("win_straat"),
        "win_postcode": record.get("win_postcode"),
        "win_plaats": record.get("win_plaats"),
        "win_land": record.get("win_land"),
        "win_contact_naam": record.get("win_contact_naam"),
        "win_contact_email": record.get("win_contact_email"),
        "win_contact_tel": record.get("win_contact_tel"),
        "win_website": record.get("win_website"),
        "buyer_bedrijf_naam": record.get("buyer_bedrijf_naam"),
        "buyer_kvk": record.get("buyer_kvk"),
        "buyer_straat": record.get("buyer_straat"),
        "buyer_postcode": record.get("buyer_postcode"),
        "buyer_plaats": record.get("buyer_plaats"),
        "buyer_land": record.get("buyer_land"),
        "buyer_contact_naam": record.get("buyer_contact_naam"),
        "buyer_contact_email": record.get("buyer_contact_email"),
        "buyer_contact_tel": record.get("buyer_contact_tel"),
        "buyer_website": record.get("buyer_website"),
        "bedrag": record.get("bedrag"),
        "valuta": record.get("valuta"),
        "publicatie_datum": pub_date,  # ← Use the variable
        "Publicatiedatum": pub_date,    # ← Use the variable
        "heeft_eerdere_aanbestedingen": record.get("heeft_eerdere_aanbestedingen", False),
        "aantal_eerdere_aanbestedingen": record.get("aantal_eerdere_aanbestedingen", 0),
    }
    if include_cpv:
        codes = [str(code) for code in (record.get("cpv_codes") or [])]
        # The legacy CPV cache uses text columns; populate both consistently
        # until all readers have moved to tender_notice_cpvs.
        mapped["cpv_codes"] = ",".join(codes)
        mapped["cpv_code"] = codes[0] if codes else None
        mapped["cpv_label"] = None
    return mapped


def map_cache_to_import_schema(cache_row: dict) -> dict:
    """
    Map een cache record terug naar het import schema.
    Inclusief historische velden.
    """
    # Extract URL - handle both dict and string formats
    url_value = cache_row.get("url") or cache_row.get("URL")
    if isinstance(url_value, dict):
        url_value = url_value.get("href") or url_value.get("url")
    
    return {
        "notice_id": cache_row.get("notice_id"),
        "publicatieId": cache_row.get("publicatie_id"),
        "publicatie_id": cache_row.get("publicatie_id"),
        "URL": url_value,
        "url": url_value,
        "titel": cache_row.get("titel"),
        "omschrijving": cache_row.get("omschrijving"),
        "win_bedrijf_naam": cache_row.get("win_bedrijf_naam"),
        "win_kvk": cache_row.get("win_kvk"),
        "win_straat": cache_row.get("win_straat"),
        "win_postcode": cache_row.get("win_postcode"),
        "win_plaats": cache_row.get("win_plaats"),
        "win_land": cache_row.get("win_land"),
        "win_contact_naam": cache_row.get("win_contact_naam"),
        "win_contact_email": cache_row.get("win_contact_email"),
        "win_contact_tel": cache_row.get("win_contact_tel"),
        "win_website": cache_row.get("win_website"),
        "buyer_bedrijf_naam": cache_row.get("buyer_bedrijf_naam"),
        "buyer_kvk": cache_row.get("buyer_kvk"),
        "buyer_straat": cache_row.get("buyer_straat"),
        "buyer_postcode": cache_row.get("buyer_postcode"),
        "buyer_plaats": cache_row.get("buyer_plaats"),
        "buyer_land": cache_row.get("buyer_land"),
        "buyer_contact_naam": cache_row.get("buyer_contact_naam"),
        "buyer_contact_email": cache_row.get("buyer_contact_email"),
        "buyer_contact_tel": cache_row.get("buyer_contact_tel"),
        "buyer_website": cache_row.get("buyer_website"),
        "bedrag": cache_row.get("bedrag"),
        "valuta": cache_row.get("valuta"),
        "contract_issue_date": cache_row.get("publicatie_datum"),
        "publicatie_datum": cache_row.get("publicatie_datum"),
        "heeft_eerdere_aanbestedingen": cache_row.get("heeft_eerdere_aanbestedingen", False),
        "aantal_eerdere_aanbestedingen": cache_row.get("aantal_eerdere_aanbestedingen", 0),
    }
def fetch_cached_in_batches(
    start_date: str,
    end_date: str,
    cache_table: str,
    cpv_codes: Optional[list] = None,
    db: Session = None,
    batch_size: int = 1000
) -> list:
    """
    Haal cached records op in batches.
    """
    CacheModel = TendernedRawCPVCached if cache_table == "tenderned_raw_cpv_cached" else TendernedRawCached
    
    # First, check total records in date range WITHOUT CPV filter
    total_in_range = db.query(func.count(CacheModel.id)).filter(
        and_(
            CacheModel.Publicatiedatum >= start_date,
            CacheModel.Publicatiedatum <= end_date
        )
    ).scalar()
    print(f"📊 Total records in date range (before CPV filter): {total_in_range}")
    
    query = db.query(CacheModel).filter(
        and_(
            CacheModel.Publicatiedatum >= start_date,
            CacheModel.Publicatiedatum <= end_date
        )
    )
    
    # Apply CPV filter if provided
    if cpv_codes and cache_table == "tenderned_raw_cpv_cached":
        print(f"🔍 Filtering by CPV codes: {cpv_codes}")
        
        # Check BOTH cpv_codes and cpv_code columns
        cpv_codes_count = db.query(func.count(CacheModel.id)).filter(
            and_(
                CacheModel.Publicatiedatum >= start_date,
                CacheModel.Publicatiedatum <= end_date,
                CacheModel.cpv_codes.isnot(None),
                CacheModel.cpv_codes != ''
            )
        ).scalar()
        
        cpv_code_count = db.query(func.count(CacheModel.id)).filter(
            and_(
                CacheModel.Publicatiedatum >= start_date,
                CacheModel.Publicatiedatum <= end_date,
                CacheModel.cpv_code.isnot(None),
                CacheModel.cpv_code != ''
            )
        ).scalar()
        
        print(f"📊 Records with non-null cpv_codes (plural): {cpv_codes_count}")
        print(f"📊 Records with non-null cpv_code (singular): {cpv_code_count}")
        
        # Sample from cpv_code column
       
        
        # Use cpv_code column instead of cpv_codes
        cpv_filters = []
        for code in cpv_codes:
            # Try matching against cpv_code column
            cpv_filters.append(
                or_(
                    CacheModel.cpv_code.like(f'%{code}%'),
                    CacheModel.cpv_codes.like(f'%{code}%')  # Also try cpv_codes just in case
                )
            )
        
        if cpv_filters:
            query = query.filter(or_(*cpv_filters))
            
            # Check count after filter
            filtered_count = query.count()
            print(f"📊 Records after CPV filter: {filtered_count}")
    
    query = query.order_by(CacheModel.id)
    
    all_records = []
    offset = 0
    
    while True:
        batch = query.limit(batch_size).offset(offset).all()
        if not batch:
            break
        
        all_records.extend([
            {
                "notice_id": record.notice_id,
                "publicatie_id": record.publicatie_id,
                "URL": record.URL_TenderNed,
                "titel": record.titel,
                "omschrijving": record.omschrijving,
                "win_bedrijf_naam": record.win_bedrijf_naam,
                "win_kvk": record.win_kvk,
                "win_straat": record.win_straat,
                "win_postcode": record.win_postcode,
                "win_plaats": record.win_plaats,
                "win_land": record.win_land,
                "win_contact_naam": record.win_contact_naam,
                "win_contact_email": record.win_contact_email,
                "win_contact_tel": record.win_contact_tel,
                "win_website": record.win_website,
                "buyer_bedrijf_naam": record.buyer_bedrijf_naam,
                "buyer_kvk": record.buyer_kvk,
                "buyer_straat": record.buyer_straat,
                "buyer_postcode": record.buyer_postcode,
                "buyer_plaats": record.buyer_plaats,
                "buyer_land": record.buyer_land,
                "buyer_contact_naam": record.buyer_contact_naam,
                "buyer_contact_email": record.buyer_contact_email,
                "buyer_contact_tel": record.buyer_contact_tel,
                "buyer_website": record.buyer_website,
                "bedrag": record.bedrag,
                "valuta": record.valuta,
                "Publicatiedatum": record.Publicatiedatum,
                "heeft_eerdere_aanbestedingen": record.heeft_eerdere_aanbestedingen,
                "aantal_eerdere_aanbestedingen": record.aantal_eerdere_aanbestedingen,
            }
            for record in batch
        ])
        
        offset += batch_size
    
    return all_records

def _run_import_job(payload: ImportRequest, user_code: str, import_id: str) -> None:
    """Run an import outside the request lifecycle, using a worker-owned session."""
    db = SessionLocal()
    try:
        import_obj = db.query(Import).filter(Import.id == import_id).first()
        if not import_obj:
            print(f"❌ Queued import {import_id} no longer exists")
            return
        _execute_import(payload, user_code, db, import_obj)
    except Exception as exc:
        # _execute_import normally records its own failure. This is a final guard
        # for errors before it can do so (for example, a connection failure).
        db.rollback()
        try:
            import_obj = db.query(Import).filter(Import.id == import_id).first()
            if import_obj and import_obj.status in ("pending", "running"):
                import_obj.status = "failed"
                import_obj.error_message = f"{type(exc).__name__}: {exc}"[:2000]
                import_obj.finished_at = dt.datetime.now(dt.timezone.utc)
                db.commit()
        except Exception:
            db.rollback()
        print(f"❌ Background import {import_id} FAILED: {exc}")
    finally:
        db.close()


@app.post("/imports", response_model=ImportJobResponse, status_code=status.HTTP_202_ACCEPTED)
def start_import(
    payload: ImportRequest,
    background_tasks: BackgroundTasks,
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Persist an import job and return immediately. GET /imports exposes its progress.

    The actual TenderNed work uses a fresh session in _run_import_job, rather
    than the request-scoped session FastAPI closes after this response.
    """
    name = generate_import_name()
    import_obj = Import(
        name=name,
        date_from=payload.date_from,
        date_to=payload.date_to,
        publicatie_type=payload.publicatie_type,
        cpv_codes=payload.cpv_codes,
        region=payload.region,
        owner_code=user_code,
        status="pending",
    )
    db.add(import_obj)
    db.commit()
    db.refresh(import_obj)

    background_tasks.add_task(_run_import_job, payload, user_code, str(import_obj.id))
    return ImportJobResponse(
        import_id=str(import_obj.id),
        name=import_obj.name,
        status=import_obj.status,
        created_at=import_obj.created_at.isoformat() if import_obj.created_at else None,
    )


def _execute_import(
    payload: ImportRequest,
    user_code: str,
    db: Session,
    import_obj: Import,
):
    """Execute a persisted import job. Called only by _run_import_job."""
    from sqlalchemy.dialects.postgresql import insert
    import traceback
    
    UPSERT_BATCH_SIZE = 1000
    
    name = import_obj.name
    import_id = import_obj.id

    # Per-run fetch counters, populated in place by run_import. Any of the four
    # scenario branches may call it; they all share this dict so the totals are
    # cumulative across the branch(es) actually taken.
    fetch_stats = {"listed": 0, "fetched": 0, "http_failed": 0, "parse_failed": 0}

    # Mark the run as in-flight. The imports row is committed before any data
    # is fetched, so without this a crash mid-import leaves a row that looks
    # identical to a successful one.
    import_obj.status = 'running'
    import_obj.started_at = dt.datetime.now(dt.timezone.utc)
    db.commit()

    try:

        # 2) Bepaal data-bron(nen): cache (met of zonder CPV) + evt run_import
        date_from = payload.date_from
        date_to = payload.date_to

        # Normaliseer naar strings voor vergelijking
        df_str = str(date_from) if date_from is not None else None
        dt_str = str(date_to) if date_to is not None else None

        records: list[dict] = []
        records_from_api: list[dict] = []

        # Bepaal welke cache table te gebruiken
        cache_table = "tenderned_raw_cpv_cached" if payload.cpv_codes else "tenderned_raw_cached"
        CacheModel = TendernedRawCPVCached if payload.cpv_codes else TendernedRawCached
        print(f"🗄️  Gebruikte cache table: {cache_table}")

        # Haal cache range op
        earliest_publicatie = db.query(func.min(CacheModel.Publicatiedatum)) \
            .filter(CacheModel.Publicatiedatum.isnot(None)) \
            .scalar()
    
        latest_publicatie = db.query(func.max(CacheModel.Publicatiedatum)) \
            .filter(CacheModel.Publicatiedatum.isnot(None)) \
            .scalar()

        ep_str = str(earliest_publicatie) if earliest_publicatie is not None else None
        lp_str = str(latest_publicatie) if latest_publicatie is not None else None

        print(f"📊 Cache range: {ep_str} tot {lp_str}")
        print(f"🎯 Gevraagde range: {df_str} tot {dt_str}")

        # SCENARIO 1: Geen cache OF startdatum > meest recente datum in cache
        if not lp_str or (df_str is not None and df_str > lp_str):
            print(f"📡 SCENARIO 1: Startdatum ({df_str}) > laatste cache ({lp_str})")
            print(f"   → Ophalen via TenderNed API: {df_str} tot {dt_str}")
        
            new_records = run_import(
                date_from=payload.date_from,
                date_to=payload.date_to,
                publicatie_type=payload.publicatie_type,
                cpv_codes=payload.cpv_codes,
                max_pages=payload.max_pages,
                stats=fetch_stats,
            )
        
            if new_records:
                records.extend(new_records)
                records_from_api.extend(new_records)
            
                # Voeg toe aan cache (bulk upsert)
                cache_records = [map_import_to_cache_schema(r, include_cpv=bool(payload.cpv_codes)) for r in new_records]
                print(f"💾 Opslaan {len(cache_records)} records in {cache_table}")
            
                try:
                    for cache_rec in cache_records:
                        stmt = insert(CacheModel).values(**cache_rec)
                        stmt = stmt.on_conflict_do_update(
                            index_elements=['notice_id'],
                            set_=cache_rec
                        )
                        db.execute(stmt)
                    db.commit()
                    print(f"✅ Cache insert successful!")
                except Exception as e:
                    print(f"❌ ERROR inserting to cache: {e}")
                    print(f"❌ Error type: {type(e)}")
                    print(f"❌ Traceback: {traceback.format_exc()}")
                    db.rollback()

        # SCENARIO 5: Startdatum < kleinste datum in cache
        elif ep_str and df_str is not None and df_str < ep_str:
            print(f"📡 SCENARIO 5: Startdatum ({df_str}) < oudste cache ({ep_str})")
        
            cache_overlap_start = ep_str
            cache_overlap_end = dt_str if (dt_str and dt_str <= lp_str) else lp_str
        
            # Haal historische data op VIA API
            api_end = ep_str
            print(f"   → API ophalen (historisch): {df_str} tot {api_end}")
        
            historical_records = run_import(
                date_from=payload.date_from,
                date_to=earliest_publicatie,
                publicatie_type=payload.publicatie_type,
                cpv_codes=payload.cpv_codes,
                max_pages=payload.max_pages,
                stats=fetch_stats,
            )
        
            if historical_records:
                records.extend(historical_records)
                records_from_api.extend(historical_records)
            
                # Voeg historische data toe aan cache
                cache_records = [map_import_to_cache_schema(r, include_cpv=bool(payload.cpv_codes)) for r in historical_records]
                print(f"💾 Opslaan {len(cache_records)} historische records in {cache_table}")
            
                try:
                    for cache_rec in cache_records:
                        stmt = insert(CacheModel).values(**cache_rec)
                        stmt = stmt.on_conflict_do_update(
                            index_elements=['notice_id'],
                            set_=cache_rec
                        )
                        db.execute(stmt)
                    db.commit()
                    print(f"✅ Cache insert successful!")
                except Exception as e:
                    print(f"❌ ERROR inserting to cache: {e}")
                    print(f"❌ Error type: {type(e)}")
                    print(f"❌ Traceback: {traceback.format_exc()}")
                    db.rollback()
        
            # Haal cached data op
            if dt_str and dt_str >= cache_overlap_start:
                print(f"   → Cache gebruiken: {cache_overlap_start} tot {cache_overlap_end}")
            
                cached_rows = fetch_cached_in_batches(
                    cache_overlap_start, 
                    cache_overlap_end,
                    cache_table=cache_table,
                    cpv_codes=payload.cpv_codes,
                    db=db
                )

                if cached_rows:
                    print(f"✅ {len(cached_rows)} records uit cache (batched)")
                    for cache_row in cached_rows:
                        records.append(map_cache_to_import_schema(cache_row))
        
            # Als einddatum > laatste cache datum, haal resterende data op
            if dt_str and lp_str and dt_str > lp_str:
                print(f"   → API ophalen (recent): {lp_str} tot {dt_str}")
            
                recent_records = run_import(
                    date_from=latest_publicatie,
                    date_to=payload.date_to,
                    publicatie_type=payload.publicatie_type,
                    cpv_codes=payload.cpv_codes,
                    max_pages=payload.max_pages,
                    stats=fetch_stats,
                )
            
                if recent_records:
                    records.extend(recent_records)
                    records_from_api.extend(recent_records)
                
                    cache_records = [map_import_to_cache_schema(r, include_cpv=bool(payload.cpv_codes)) for r in recent_records]
                    print(f"💾 Opslaan {len(cache_records)} recente records in {cache_table}")
                
                    try:
                        for cache_rec in cache_records:
                            stmt = insert(CacheModel).values(**cache_rec)
                            stmt = stmt.on_conflict_do_update(
                                index_elements=['notice_id'],
                                set_=cache_rec
                            )
                            db.execute(stmt)
                        db.commit()
                        print(f"✅ Cache insert successful!")
                    except Exception as e:
                        print(f"❌ ERROR inserting to cache: {e}")
                        print(f"❌ Error type: {type(e)}")
                        print(f"❌ Traceback: {traceback.format_exc()}")
                        db.rollback()

        # SCENARIO 2 & 3: Startdatum <= meest recente datum in cache
        else:
            print(f"📦 SCENARIO 2/3: Startdatum ({df_str}) binnen cache range")
        
            upper_bound_str = dt_str if (dt_str is not None and dt_str <= lp_str) else lp_str
        
            print(f"   → Cache gebruiken: {df_str} tot {upper_bound_str}")
            cached_rows = fetch_cached_in_batches(
                df_str, 
                upper_bound_str,
                cache_table=cache_table,
                cpv_codes=payload.cpv_codes,
                db=db
            )

            if cached_rows:
                print(f"✅ {len(cached_rows)} records uit cache (batched)")
                for cache_row in cached_rows:
                    records.append(map_cache_to_import_schema(cache_row))

            # Als einddatum > laatste cache datum, vul aan met API
            if dt_str is not None and lp_str and dt_str > lp_str:
                print(f"   → API aanvullen: {lp_str} tot {dt_str}")
            
                new_records = run_import(
                    date_from=latest_publicatie,
                    date_to=payload.date_to,
                    publicatie_type=payload.publicatie_type,
                    cpv_codes=payload.cpv_codes,
                    max_pages=payload.max_pages,
                    stats=fetch_stats,
                )

                if new_records:
                    print(f"✅ {len(new_records)} nieuwe records via API")
                    records.extend(new_records)
                    records_from_api.extend(new_records)

                    # Voeg toe aan cache
                    cache_records = [map_import_to_cache_schema(r, include_cpv=bool(payload.cpv_codes)) for r in new_records]
                    print(f"💾 Opslaan {len(cache_records)} nieuwe records in {cache_table}")
                
                    try:
                        for cache_rec in cache_records:
                            stmt = insert(CacheModel).values(**cache_rec)
                            stmt = stmt.on_conflict_do_update(
                                index_elements=['notice_id'],
                                set_=cache_rec
                            )
                            db.execute(stmt)
                        db.commit()
                        print(f"✅ Cache insert successful!")
                    except Exception as e:
                        print(f"❌ ERROR inserting to cache: {e}")
                        print(f"❌ Error type: {type(e)}")
                        print(f"❌ Traceback: {traceback.format_exc()}")
                        db.rollback()

        print(f"📊 Totaal records verzameld: {len(records)}")
        print(f"📡 Waarvan van API: {len(records_from_api)}")

        # Persist the immutable source XML/listing before producing the
        # owner-scoped import projection. If this fails, the job fails rather
        # than claiming a cacheable import whose source cannot be reproduced.
        persist_canonical_records(records_from_api, db)

        # Maak een set van notice_ids die van API komen
        api_notice_ids = {r.get("notice_id") for r in records_from_api if r.get("notice_id")}

        # 3) Map records & voeg import_id toe
        notice_rows = []
        for r in records:
            province = province_for_record(r)
        
            # Skip dit record als region filter actief is EN province niet matcht
            if payload.region:
                if not province:
                    continue
                if province.strip().lower() != payload.region.strip().lower():
                    continue

            notice_id = r.get("notice_id")
        
            # Check historische aanbestedingen ALLEEN als dit record van API komt
            heeft_eerdere_aanbestedingen = r.get("heeft_eerdere_aanbestedingen", False)
            aantal_eerdere_aanbestedingen = r.get("aantal_eerdere_aanbestedingen", 0)
        
            if notice_id in api_notice_ids:
                bedrijf_naam = (r.get("win_bedrijf_naam") or "").strip()
            
                if bedrijf_naam:
                    try:
                        historische_records = _search_companies_in_db(
                            q=bedrijf_naam,
                            years=5,
                            max_results=100,
                            db=db
                        )
                    
                        if historische_records:
                            heeft_eerdere_aanbestedingen = True
                            aantal_eerdere_aanbestedingen = len(historische_records)
                    except Exception as e:
                        print(f"Fout bij zoeken historische data voor {bedrijf_naam}: {e}")

            # Haal publicatie_datum op
            pub_datum = r.get("contract_issue_date") or r.get("publicatie_datum")
            if pub_datum and isinstance(pub_datum, str) and len(pub_datum) >= 10:
                pub_datum = pub_datum[:10]
        
            # Extract URL - handle both dict and string formats
            url_value = r.get("URL") or r.get("url")
            if isinstance(url_value, dict):
                url_value = url_value.get("href") or url_value.get("url")
        
            notice_rows.append({
                "import_id": import_id,
                "notice_id": notice_id,
                "publicatie_id": r.get("publicatieId") or r.get("publicatie_id"),
                "url": url_value,
                "titel": r.get("titel"),
                "omschrijving": r.get("omschrijving"),
                "win_bedrijf_naam": r.get("win_bedrijf_naam"),
                "win_kvk": r.get("win_kvk"),
                "win_straat": r.get("win_straat"),
                "win_postcode": r.get("win_postcode"),
                "win_plaats": r.get("win_plaats"),
                "win_land": r.get("win_land"),
                "win_contact_naam": r.get("win_contact_naam"),
                "win_contact_email": r.get("win_contact_email"),
                "win_contact_tel": r.get("win_contact_tel"),
                "win_website": r.get("win_website"),
                "buyer_bedrijf_naam": r.get("buyer_bedrijf_naam"),
                "buyer_kvk": r.get("buyer_kvk"),
                "buyer_straat": r.get("buyer_straat"),
                "buyer_postcode": r.get("buyer_postcode"),
                "buyer_plaats": r.get("buyer_plaats"),
                "buyer_land": r.get("buyer_land"),
                "buyer_contact_naam": r.get("buyer_contact_naam"),
                "buyer_contact_email": r.get("buyer_contact_email"),
                "buyer_contact_tel": r.get("buyer_contact_tel"),
                "buyer_website": r.get("buyer_website"),
                "bedrag": r.get("bedrag"),
                "valuta": r.get("valuta"),
                "province": province,
                "publicatie_datum": pub_datum,
                "heeft_eerdere_aanbestedingen": heeft_eerdere_aanbestedingen,
                "aantal_eerdere_aanbestedingen": aantal_eerdere_aanbestedingen,
                "owner_code": user_code,
                # Classification: filter on record_type == 'awards' to exclude
                # VEAT (no winner yet) and cancelled procedures. See migration 004.
                "publicatie_code": r.get("publicatie_code"),
                "record_type": r.get("record_type"),
                "type_publicatie": r.get("type_publicatie"),
                "is_cancelled": bool(r.get("is_cancelled")),
            })

        print(f"✅ {len(notice_rows)} records na filtering")

        # 4) Deduplicate and bulk upsert naar database
        unique_rows = {}
        for row in notice_rows:
            nid = row.get("notice_id")
            if not nid:
                continue
            unique_rows[nid] = row

        deduped_notice_rows = list(unique_rows.values())
        total_records = len(deduped_notice_rows)
        print(f"💾 Opslaan {total_records} unieke notices (van {len(notice_rows)} totaal)")

        if deduped_notice_rows:
            total_batches = (len(deduped_notice_rows) - 1) // UPSERT_BATCH_SIZE + 1
        
            for i in range(0, len(deduped_notice_rows), UPSERT_BATCH_SIZE):
                batch = deduped_notice_rows[i:i + UPSERT_BATCH_SIZE]
                batch_num = i // UPSERT_BATCH_SIZE + 1
                print(f"💾 Upserting batch {batch_num}/{total_batches} ({len(batch)} records)")
            
                # Bulk upsert using PostgreSQL's ON CONFLICT
                stmt = insert(Notice).values(batch)
                # Conflict target is (import_id, notice_id), NOT notice_id alone.
                # With a global notice_id key, an overlapping import re-parented another
                # user's rows to itself and overwrote their owner_code, silently emptying
                # the earlier import. See migrations/001_fix_notice_uniqueness.sql.
                stmt = stmt.on_conflict_do_update(
                    index_elements=['import_id', 'notice_id'],
                    set_={
                        'publicatie_id': stmt.excluded.publicatie_id,
                        'url': stmt.excluded.url,
                        'titel': stmt.excluded.titel,
                        'omschrijving': stmt.excluded.omschrijving,
                        'win_bedrijf_naam': stmt.excluded.win_bedrijf_naam,
                        'win_kvk': stmt.excluded.win_kvk,
                        'win_straat': stmt.excluded.win_straat,
                        'win_postcode': stmt.excluded.win_postcode,
                        'win_plaats': stmt.excluded.win_plaats,
                        'win_land': stmt.excluded.win_land,
                        'win_contact_naam': stmt.excluded.win_contact_naam,
                        'win_contact_email': stmt.excluded.win_contact_email,
                        'win_contact_tel': stmt.excluded.win_contact_tel,
                        'win_website': stmt.excluded.win_website,
                        'buyer_bedrijf_naam': stmt.excluded.buyer_bedrijf_naam,
                        'buyer_kvk': stmt.excluded.buyer_kvk,
                        'buyer_straat': stmt.excluded.buyer_straat,
                        'buyer_postcode': stmt.excluded.buyer_postcode,
                        'buyer_plaats': stmt.excluded.buyer_plaats,
                        'buyer_land': stmt.excluded.buyer_land,
                        'buyer_contact_naam': stmt.excluded.buyer_contact_naam,
                        'buyer_contact_email': stmt.excluded.buyer_contact_email,
                        'buyer_contact_tel': stmt.excluded.buyer_contact_tel,
                        'buyer_website': stmt.excluded.buyer_website,
                        'bedrag': stmt.excluded.bedrag,
                        'valuta': stmt.excluded.valuta,
                        'province': stmt.excluded.province,
                        'publicatie_datum': stmt.excluded.publicatie_datum,
                        'publicatie_code': stmt.excluded.publicatie_code,
                    'record_type': stmt.excluded.record_type,
                    'type_publicatie': stmt.excluded.type_publicatie,
                    'is_cancelled': stmt.excluded.is_cancelled,
                    'heeft_eerdere_aanbestedingen': stmt.excluded.heeft_eerdere_aanbestedingen,
                        'aantal_eerdere_aanbestedingen': stmt.excluded.aantal_eerdere_aanbestedingen,
                        'owner_code': stmt.excluded.owner_code,
                    }
                )
            
                try:
                    db.execute(stmt)
                    db.commit()
                except Exception as e:
                    print(f"❌ Error batch {batch_num}: {e}")
                    db.rollback()
                    raise

        # 5) Update total_records
        import_obj.total_records = total_records
        db.commit()

        dropped = fetch_stats["http_failed"] + fetch_stats["parse_failed"]
        print(f"✅ Import {name} voltooid: {total_records} records")
        if dropped:
            print(
                f"⚠️ Import {name} INCOMPLEET: {dropped} van {fetch_stats['listed']} "
                f"publicaties niet verwerkt "
                f"(http_failed={fetch_stats['http_failed']} parse_failed={fetch_stats['parse_failed']})"
            )

        # 'partial' means the run finished but did not store everything the
        # list endpoint advertised -- previously indistinguishable from success.
        import_obj.status = 'partial' if dropped else 'completed'
        import_obj.finished_at = dt.datetime.now(dt.timezone.utc)
        import_obj.listed = fetch_stats["listed"]
        import_obj.fetched = fetch_stats["fetched"]
        import_obj.http_failed = fetch_stats["http_failed"]
        import_obj.parse_failed = fetch_stats["parse_failed"]
        db.commit()

        return ImportResponse(
            import_id=str(import_id),
            name=name,
            total_records=total_records,
            created_at=import_obj.created_at.isoformat() if import_obj.created_at else None,
            listed=fetch_stats["listed"],
            fetched=fetch_stats["fetched"],
            http_failed=fetch_stats["http_failed"],
            parse_failed=fetch_stats["parse_failed"],
            complete=(dropped == 0),
        )
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        import_obj.status = 'failed'
        import_obj.error_message = f'{type(exc).__name__}: {exc}'[:2000]
        import_obj.finished_at = dt.datetime.now(dt.timezone.utc)
        db.commit()
        print(f'❌ Import {name} FAILED: {exc}')
        raise

@app.get("/imports/{import_id}/notices")
def get_import_notices(
    import_id: str, 
    region: Optional[str] = Query(None, description="Filter op provincie/regio"), 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Fetch all notices for a specific import to display in the UI.
    """
    print("\n=== FETCH NOTICES START ===")
    print(f"Import ID received: {import_id!r}")

    try:
        # Query notices using SQLAlchemy
        query = db.query(Notice).filter(
            Notice.import_id == import_id,
            Notice.owner_code == user_code
        ).order_by(Notice.created_at.asc())

        # Apply region filter if provided
        if region:
            region_norm = region.strip().lower()
            query = query.filter(
                func.lower(Notice.province) == region_norm
            )

        notices = query.all()
        print(f"SQLAlchemy returned {len(notices)} notices for import_id={import_id!r}")

        # Convert SQLAlchemy objects to dicts
        rows = []
        for notice in notices:
            notice_dict = {
                "id": str(notice.id) if notice.id else None,
                "import_id": str(notice.import_id) if notice.import_id else None,
                "notice_id": notice.notice_id,
                "publicatie_id": notice.publicatie_id,
                "url": notice.url,
                "titel": notice.titel,
                "omschrijving": notice.omschrijving,
                "win_bedrijf_naam": notice.win_bedrijf_naam,
                "win_kvk": notice.win_kvk,
                "win_straat": notice.win_straat,
                "win_postcode": notice.win_postcode,
                "win_plaats": notice.win_plaats,
                "win_land": notice.win_land,
                "win_contact_naam": notice.win_contact_naam,
                "win_contact_email": notice.win_contact_email,
                "win_contact_tel": notice.win_contact_tel,
                "win_website": notice.win_website,
                "buyer_bedrijf_naam": notice.buyer_bedrijf_naam,
                "buyer_kvk": notice.buyer_kvk,
                "buyer_straat": notice.buyer_straat,
                "buyer_postcode": notice.buyer_postcode,
                "buyer_plaats": notice.buyer_plaats,
                "buyer_land": notice.buyer_land,
                "buyer_contact_naam": notice.buyer_contact_naam,
                "buyer_contact_email": notice.buyer_contact_email,
                "buyer_contact_tel": notice.buyer_contact_tel,
                "buyer_website": notice.buyer_website,
                "bedrag": float(notice.bedrag) if notice.bedrag else None,
                "valuta": notice.valuta,
                "province": notice.province,
                "publicatie_datum": notice.publicatie_datum,
                "heeft_eerdere_aanbestedingen": notice.heeft_eerdere_aanbestedingen,
                "aantal_eerdere_aanbestedingen": notice.aantal_eerdere_aanbestedingen,
                "created_at": notice.created_at.isoformat() if notice.created_at else None,
            }
            rows.append(notice_dict)

    except Exception as e:
        print(f"ERROR while fetching notices for import_id={import_id!r}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    print("=== FETCH NOTICES END ===\n")
    return rows


@app.get("/imports/{import_id}/download")
def download_import(
    import_id: str,
    format: str = Query("excel", pattern="^(excel|csv)$"),
    region: Optional[str] = Query(None, description="Filter op provincie/regio voor download"),
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Genereer Excel of CSV op basis van SQLAlchemy-data.
    """
    # Query notices using SQLAlchemy
    query = db.query(Notice).filter(
        Notice.import_id == import_id,
        Notice.owner_code == user_code
    )

    # Apply region filter if requested
    if region:
        region_norm = region.strip().lower()
        query = query.filter(
            func.lower(Notice.province) == region_norm
        )

    notices = query.all()
    
    if not notices:
        raise HTTPException(404, "Geen records voor deze import")

    # Convert SQLAlchemy objects to dicts
    rows = []
    for notice in notices:
        notice_dict = {
            "id": str(notice.id) if notice.id else None,
            "import_id": str(notice.import_id) if notice.import_id else None,
            "notice_id": notice.notice_id,
            "publicatie_id": notice.publicatie_id,
            "url": notice.url,
            "titel": notice.titel,
            "omschrijving": notice.omschrijving,
            "win_bedrijf_naam": notice.win_bedrijf_naam,
            "win_kvk": notice.win_kvk,
            "win_straat": notice.win_straat,
            "win_postcode": notice.win_postcode,
            "win_plaats": notice.win_plaats,
            "win_land": notice.win_land,
            "win_contact_naam": notice.win_contact_naam,
            "win_contact_email": notice.win_contact_email,
            "win_contact_tel": notice.win_contact_tel,
            "win_website": notice.win_website,
            "buyer_bedrijf_naam": notice.buyer_bedrijf_naam,
            "buyer_kvk": notice.buyer_kvk,
            "buyer_straat": notice.buyer_straat,
            "buyer_postcode": notice.buyer_postcode,
            "buyer_plaats": notice.buyer_plaats,
            "buyer_land": notice.buyer_land,
            "buyer_contact_naam": notice.buyer_contact_naam,
            "buyer_contact_email": notice.buyer_contact_email,
            "buyer_contact_tel": notice.buyer_contact_tel,
            "buyer_website": notice.buyer_website,
            "bedrag": float(notice.bedrag) if notice.bedrag else None,
            "valuta": notice.valuta,
            "province": notice.province,
            "publicatie_datum": notice.publicatie_datum,
            "heeft_eerdere_aanbestedingen": notice.heeft_eerdere_aanbestedingen,
            "aantal_eerdere_aanbestedingen": notice.aantal_eerdere_aanbestedingen,
            "created_at": notice.created_at.isoformat() if notice.created_at else None,
        }
        rows.append(notice_dict)

    if format == "csv":
        return _make_csv_response(rows, filename=f"import_{import_id}.csv")
    else:
        return _make_excel_response(rows, filename=f"import_{import_id}.xlsx")


def _make_csv_response(rows, filename: str):
    import csv

    output = io.StringIO()
    writer = None
    for row in rows:
        if writer is None:
            fieldnames = list(row.keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _excel_safe(value):
    """Make a value writable by openpyxl.

    Excel has no concept of timezones, so openpyxl raises TypeError on any
    tz-aware datetime/time. Our Postgres columns are TIMESTAMPTZ, so strip the
    tzinfo (after converting to local time) before writing.
    """
    if isinstance(value, dt.datetime) and value.tzinfo is not None:
        return value.astimezone().replace(tzinfo=None)
    if isinstance(value, dt.time) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _make_excel_response(rows, filename: str):
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "notices"

    if not rows:
        pass
    else:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([_excel_safe(row.get(h)) for h in headers])

        # Optioneel: auto column width
        for i, col in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


from datetime import datetime, timedelta, date
from decimal import Decimal

def _search_companies_in_db(q: str, years: int = 5, max_results: int = 1000, db: Session = None) -> List[Dict]:
    """
    Doorzoek de tabel tenderned_raw op bedrijfsnaam (case-insensitive substring)
    en filter op Publicatiedatum binnen de afgelopen `years` jaren.
    Dedupliceert op basis van bedrijf + KVK + omschrijving en behoudt alleen
    de meest recente publicatie.
    Extraheert begin- en einddatum van de aanbesteding.

    Return-format (nu met bedrag) sluit aan op de frontend:

    [
      {
        'publicatiedatum': '2024-03-10',
        'year': '2024',
        'bedrijf': 'Studytube B.V.',
        'kvk': '51290901',
        'omschrijving': '...',
        'begindatum_opdracht': '2024-12-01',
        'einddatum_opdracht': '2029-02-28',
        'aantal_publicaties': 3,
        'bedrag': 123456.78,
      },
      ...
    ]
    """
    if not q:
        return []
    
    if not db:
        raise ValueError("Database session is required")

    cutoff = (datetime.utcnow() - timedelta(days=365 * int(years))).date()

    # The canonical layer is the source of truth for new imports. Keep the
    # legacy table fallback only until historic data has been backfilled.
    try:
        canonical_rows = db.query(TenderNotice, TenderCompany).join(
            TenderCompany, TenderNotice.winner_company_id == TenderCompany.id
        ).filter(
            TenderCompany.name.ilike(f"%{q}%"),
            TenderNotice.publicatie_datum >= cutoff,
        ).order_by(TenderNotice.publicatie_datum.desc()).limit(max_results).all()
    except Exception as e:
        print(f"Database query error: {e}")
        raise

    if canonical_rows:
        results = []
        for notice, company in canonical_rows:
            lot = db.query(TenderNoticeLot).filter(
                TenderNoticeLot.notice_id == notice.id
            ).order_by(TenderNoticeLot.award_value.desc().nullslast()).first()
            results.append({
                "publicatiedatum": str(notice.publicatie_datum) if notice.publicatie_datum else None,
                "year": str(notice.publicatie_datum.year) if notice.publicatie_datum else None,
                "bedrijf": company.name,
                "kvk": company.kvk,
                "omschrijving": notice.description,
                "begindatum_opdracht": str(lot.contract_start) if lot and lot.contract_start else None,
                "einddatum_opdracht": str(lot.contract_end) if lot and lot.contract_end else None,
                "aantal_publicaties": 1,
                "bedrag": float(lot.award_value) if lot and lot.award_value is not None else None,
            })
        return results

    # Legacy fallback for pre-migration imports.
    rows = db.query(TendernedRaw).filter(
        TendernedRaw.Officiele_benaming.ilike(f"%{q}%"),
        TendernedRaw.Publicatiedatum >= cutoff
    ).limit(max_results * 5).all()

    if not rows:
        return []

    # Helper voor datums
    def parse_date(raw):
        if not raw:
            return None
        if isinstance(raw, date):
            return datetime.combine(raw, datetime.min.time())
        if isinstance(raw, datetime):
            return raw
        if isinstance(raw, str):
            try:
                return datetime.fromisoformat(raw)
            except ValueError:
                try:
                    return datetime.strptime(raw, "%d-%m-%Y")
                except ValueError:
                    return None
        return None

    # Helper voor bedrag
    def parse_amount(raw):
        if raw is None:
            return None
        if isinstance(raw, (int, float, Decimal)):
            return float(raw)
        if isinstance(raw, str):
            txt = raw.replace(".", "").replace(",", ".") if "," in raw else raw
            try:
                return float(txt)
            except ValueError:
                return None
        return None

    # 2) Groepeer per unieke combinatie (bedrijf, kvk, omschrijving[:100])
    all_matches = defaultdict(list)

    q_lower = q.lower()
    for row in rows:
        # Use Python attribute names (without spaces)
        company = (row.Officiele_benaming or "").strip()
        if not company:
            continue

        if q_lower not in company.lower():
            continue

        kvk = (row.Kvknummer or "").strip()
        omschrijving = (row.Omschrijving_aanbesteding or "").strip()

        pub_raw = row.Publicatiedatum
        pub_date = parse_date(pub_raw)

        # bedrag uit kolom "bedrag"
        bedrag_raw = row.bedrag if hasattr(row, 'bedrag') else None
        bedrag = parse_amount(bedrag_raw)

        unique_key = (company, kvk, omschrijving[:100] if omschrijving else "")

        all_matches[unique_key].append(
            {
                "publicatiedatum": str(pub_raw) if pub_raw else None,
                "pub_date_obj": pub_date,
                "row": row,
                "company": company,
                "kvk": kvk,
                "omschrijving": omschrijving,
                "bedrag": bedrag,
            }
        )

    # 3) Tweede pass: pak per key de meest recente publicatie
    results = []

    for unique_key, matches in all_matches.items():
        if len(results) >= max_results:
            break

        # Kies de meest recente o.b.v. pub_date_obj (fallback op string)
        def sort_key(m):
            if m["pub_date_obj"] is not None:
                return m["pub_date_obj"]
            return m["publicatiedatum"] or ""

        most_recent = max(matches, key=sort_key)
        row = most_recent["row"]
        pub_date = most_recent["pub_date_obj"]

        # Begin- en einddatum uit row - use Python attribute names
        aanvang_raw = (row.Aanvang_opdracht or "") if hasattr(row, 'Aanvang_opdracht') else ""
        voltooiing_raw = (row.Voltooiing_opdracht or "") if hasattr(row, 'Voltooiing_opdracht') else ""

        def to_iso_str(raw):
            if not raw:
                return None
            d = parse_date(raw)
            return d.strftime("%Y-%m-%d") if d else str(raw)

        begindatum = to_iso_str(aanvang_raw)
        einddatum = to_iso_str(voltooiing_raw)

        results.append(
            {
                "publicatiedatum": most_recent["publicatiedatum"],
                "year": str(pub_date.year) if pub_date else "",
                "bedrijf": most_recent["company"],
                "kvk": most_recent["kvk"],
                "omschrijving": most_recent["omschrijving"],
                "begindatum_opdracht": begindatum,
                "einddatum_opdracht": einddatum,
                "aantal_publicaties": len(matches),
                "bedrag": most_recent["bedrag"],
            }
        )

    return results

    
@app.post("/validate-code")
def validate_code_endpoint(payload: dict, db: Session = Depends(get_db)):
    """Validate a code server-side without needing a full authenticated request.
    Expects JSON: { code: "123456789012" }
    Returns 200 ok if code exists in `auth_codes` table; 400 for bad input; 403 if unknown.
    """
    code = (payload.get("code") if isinstance(payload, dict) else None) or ""
    code = str(code).strip()
    if not code or not code.isdigit() or len(code) != 12:
        raise HTTPException(status_code=400, detail="Invalid code format")

    try:
        auth_code = db.query(AuthCode).filter(AuthCode.code == code).first()
        if not auth_code:
            raise HTTPException(status_code=403, detail="Unknown code")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/auth/request-code")
def request_code_endpoint(db: Session = Depends(get_db)):
    """Generate a unique 12-digit numeric code, store it in `auth_codes` and return it.
    This endpoint is intentionally simple: it returns the generated code in the response body.
    Ensure your DB has an `auth_codes(code TEXT PRIMARY KEY, created_at TIMESTAMPTZ)` table.
    """
    def _gen():
        return "%012d" % random.randint(0, 10**12 - 1)

    # Try a few times to avoid collisions on insert
    attempts = 0
    max_attempts = 8
    while attempts < max_attempts:
        attempts += 1
        code = _gen()
        try:
            # Check if code already exists
            existing = db.query(AuthCode).filter(AuthCode.code == code).first()
            if existing:
                continue
            
            # Insert new e
            new_code = AuthCode(
                code=code,
                created_at=dt.datetime.utcnow()
            )
            db.add(new_code)
            db.commit()
            
            return {"code": code}
            
        except Exception as e:
            db.rollback()
            if attempts >= max_attempts:
                raise HTTPException(status_code=500, detail=f"Error generating code: {e}")
            continue

    raise HTTPException(status_code=500, detail="Could not generate unique code, try again later")


# ========================================
# SEARCH ENDPOINT
# ========================================

@app.get("/api/search/company")
def api_search_company(
    q: str = Query(..., description="Bedrijfsnaam om te zoeken"),
    years: int = Query(5, ge=1, le=20),
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    API endpoint om op bedrijfsnaam te zoeken in TenderNed_filtered.csv.
    Query params:
      - q: bedrijfsnaam (substring, verplicht)
      - years: hoeveel jaren terug (default 5)
    
    Retourneert gededupliceerde resultaten met begin- en einddatum van opdrachten.
    """
    try:
        results = _search_companies_in_db(q=q, years=years, max_results=1000, db=db)
        print(f"Found {len(results)} results for query '{q}' within last {years} years.")
        print(results)
        return {"query": q, "years": years, "total": len(results), "results": results}
    except FileNotFoundError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")


# ========================================
# SROI ANALYSIS ENDPOINTS
# ========================================

def get_notices_for_sroi(import_id: str, owner_code: str, db: Session) -> List[Dict]:
    """
    Haal alle notices op voor SROI analyse.
    """
    notices = db.query(Notice)\
        .filter(Notice.import_id == import_id)\
        .filter(Notice.owner_code == owner_code)\
        .all()
    
    # Convert to dict format (assuming you need dict for analysis functions)
    return [notice.__dict__ for notice in notices]


def save_sroi_results(results: List[Dict], owner_code: Optional[str] = None, db: Session = None):
    """
    Sla SROI resultaten op in database.
    """
    if not results or not db:
        return
    
    for result in results:
        # Check if result already exists
        existing = db.query(SROIResult)\
            .filter(SROIResult.import_id == result.get("import_id"))\
            .filter(SROIResult.notice_id == result.get("notice_id"))\
            .first()
        
        if existing:
            # Update existing
            for key, value in result.items():
                if hasattr(existing, key):
                    setattr(existing, key, value)
            existing.owner_code = owner_code
        else:
            # Create new
            new_result = SROIResult(
                import_id=result.get("import_id"),
                notice_id=result.get("notice_id"),
                publicatie_id=result.get("publicatie_id"),
                winner_name=result.get("winner_name"),
                analyzed_url=result.get("analyzed_url"),
                url_source=result.get("url_source"),
                sroi_compliant=result.get("sroi_compliant", False),
                confidence=result.get("confidence", "none"),
                score=result.get("score", 0),
                evidence=result.get("evidence", []),
                summary=result.get("summary", ""),
                pages_checked=result.get("pages_checked", 0),
                error=result.get("error"),
                analysis_method=result.get("analysis_method"),
                verdict=result.get("verdict"),
                owner_code=owner_code
            )
            db.add(new_result)
    
    db.commit()


def run_sroi_analysis_background(import_id: str, owner_code: Optional[str] = None):
    """
    Background task voor SROI analyse.
    """
    # Create a new database session for this background task
    from database import SessionLocal
    db = SessionLocal()
    
    try:
        print(f"\n🚀 Starting SROI analysis for import {import_id}")

        # Update status
        sroi_analysis_status[import_id] = {
            "status": "running",
            "progress": 0,
            "current": 0,
            "total": 0,
            "started_at": dt.datetime.utcnow().isoformat()
        }

        # Haal notices op (eigenaar filter)
        notices = get_notices_for_sroi(import_id, owner_code, db)

        if not notices:
            sroi_analysis_status[import_id].update({
                "status": "failed",
                "error": "Geen notices gevonden voor deze import"
            })
            print(f"❌ No notices found for import {import_id}")
            return

        sroi_analysis_status[import_id]["total"] = len(notices)
        print(f"📊 Found {len(notices)} notices to analyze")

        # Progress callback
        def progress_callback(current, total, result):
            sroi_analysis_status[import_id].update({
                "current": current,
                "progress": int((current / total) * 100)
            })
            print(f"Progress: {current}/{total} ({sroi_analysis_status[import_id]['progress']}%)")

        # Run analysis
        results = analyze_import_sroi(notices, progress_callback)

        # Add import_id to each result
        for result in results:
            result["import_id"] = import_id

        # Save to database
        print(f"💾 Saving {len(results)} results to database...")
        save_sroi_results(results, owner_code=owner_code, db=db)

        # Calculate summary
        compliant_count = sum(1 for r in results if r.get("sroi_compliant"))
        avg_score = sum(r.get("score", 0) for r in results) / len(results) if results else 0

        sroi_analysis_status[import_id].update({
            "status": "completed",
            "progress": 100,
            "completed_at": dt.datetime.utcnow().isoformat(),
            "results_summary": {
                "total": len(results),
                "compliant": compliant_count,
                "non_compliant": len(results) - compliant_count,
                "compliance_rate": (compliant_count / len(results) * 100) if results else 0,
                "average_score": round(avg_score, 2)
            }
        })

        print(f"✅ SROI analysis completed for import {import_id}")
        print(f"   Compliant: {compliant_count}/{len(results)} ({compliant_count/len(results)*100:.1f}%)")
        print(f"   Average score: {avg_score:.1f}\n")

    except Exception as e:
        print(f"❌ ERROR in SROI analysis for import {import_id}: {e}")
        sroi_analysis_status[import_id].update({
            "status": "failed",
            "error": str(e)
        })
        db.rollback()
    finally:
        db.close()


@app.post("/imports/{import_id}/sroi-analyze")
def start_sroi_analysis(
    import_id: str, 
    background_tasks: BackgroundTasks, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Start SROI analyse voor een import.
    """
    # Check if import exists and belongs to caller
    import_record = db.query(Import)\
        .filter(Import.id == import_id)\
        .first()
    
    if not import_record:
        raise HTTPException(status_code=404, detail="Import niet gevonden")
    if import_record.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toestemming voor deze import")
    
    # Check if analysis already running
    if import_id in sroi_analysis_status:
        status = sroi_analysis_status[import_id]
        if status["status"] == "running":
            raise HTTPException(
                status_code=400, 
                detail="Analyse is al bezig voor deze import"
            )
    
    # Check if results already exist
    existing = db.query(SROIResult)\
        .filter(SROIResult.import_id == import_id)\
        .first()
    
    if existing:
        raise HTTPException(
            status_code=400,
            detail="Er bestaan al SROI resultaten voor deze import. Verwijder ze eerst om opnieuw te analyseren."
        )
    
    # Start background task (pass owner_code so saved results are scoped)
    background_tasks.add_task(run_sroi_analysis_background, import_id, user_code)
    
    # Initialize status
    sroi_analysis_status[import_id] = {
        "status": "pending",
        "progress": 0,
        "current": 0,
        "total": 0,
        "started_at": dt.datetime.utcnow().isoformat()
    }
    
    return {
        "message": "SROI analyse gestart",
        "import_id": import_id,
        "status": "pending"
    }


def save_notice_sroi_result(result: Dict, owner_code: str, db: Session):
    """
    Sla één SROI resultaat op voor een specifieke notice (winner of buyer).
    
    Args:
        result: Dictionary met SROI analyse resultaat
        owner_code: User code van de eigenaar
        db: Database session
    """
    if not result:
        return
    
    # Check if result already exists
    existing = db.query(NoticeSROIResult)\
        .filter(NoticeSROIResult.notice_id == result.get("notice_id"))\
        .filter(NoticeSROIResult.target == result.get("target", "winner"))\
        .first()
    
    if existing:
        # Update existing
        existing.company_name = result.get("winner_name") or result.get("company_name")
        existing.analyzed_url = result.get("analyzed_url")
        existing.url_source = result.get("url_source")
        existing.sroi_compliant = result.get("sroi_compliant", False)
        existing.confidence = result.get("confidence", "none")
        existing.score = result.get("score", 0)
        existing.evidence = result.get("evidence", [])
        existing.summary = result.get("summary", "")
        existing.pages_checked = result.get("pages_checked", 0)
        existing.error = result.get("error")
        existing.analysis_method = result.get("analysis_method")
        existing.verdict = result.get("verdict")
        existing.owner_code = owner_code
    else:
        # Create new
        new_result = NoticeSROIResult(
            notice_id=result.get("notice_id"),
            target=result.get("target", "winner"),
            company_name=result.get("winner_name") or result.get("company_name"),
            analyzed_url=result.get("analyzed_url"),
            url_source=result.get("url_source"),
            sroi_compliant=result.get("sroi_compliant", False),
            confidence=result.get("confidence", "none"),
            score=result.get("score", 0),
            evidence=result.get("evidence", []),
            summary=result.get("summary", ""),
            pages_checked=result.get("pages_checked", 0),
            error=result.get("error"),
            analysis_method=result.get("analysis_method"),
            verdict=result.get("verdict"),
            owner_code=owner_code
        )
        db.add(new_result)
    
    db.commit()


@app.get("/imports/{import_id}/sroi-status")
def get_sroi_status(
    import_id: str, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Haal de status op van een lopende SROI analyse.
    """
    # Ensure import belongs to user
    import_record = db.query(Import)\
        .filter(Import.id == import_id)\
        .first()
    
    if not import_record:
        raise HTTPException(status_code=404, detail="Import niet gevonden")
    if import_record.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toegang tot de status van deze import")

    if import_id not in sroi_analysis_status:
        # Check if results exist in database
        result = db.query(SROIResult)\
            .filter(SROIResult.import_id == import_id)\
            .first()
        
        if result:
            return {
                "import_id": import_id,
                "status": "completed",
                "message": "Analyse is al uitgevoerd. Bekijk de resultaten."
            }
        
        return {
            "import_id": import_id,
            "status": "not_started",
            "message": "Nog geen analyse gestart voor deze import"
        }
    
    return {
        "import_id": import_id,
        **sroi_analysis_status[import_id]
    }


@app.get("/imports/{import_id}/sroi-results")
def get_sroi_results(
    import_id: str, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Haal SROI resultaten op voor een import.
    """
    # Only fetch results for this owner's import
    results = db.query(SROIResult)\
        .filter(SROIResult.import_id == import_id)\
        .filter(SROIResult.owner_code == user_code)\
        .order_by(desc(SROIResult.score))\
        .all()
    
    # Convert to dict
    results_list = [result.__dict__ for result in results]
    # Remove SQLAlchemy internal state
    for r in results_list:
        r.pop('_sa_instance_state', None)
    
    # Calculate summary
    if results_list:
        compliant_count = sum(1 for r in results_list if r["sroi_compliant"])
        avg_score = sum(r["score"] for r in results_list) / len(results_list)
        
        summary = {
            "total": len(results_list),
            "compliant": compliant_count,
            "non_compliant": len(results_list) - compliant_count,
            "compliance_rate": round((compliant_count / len(results_list) * 100), 2),
            "average_score": round(avg_score, 2)
        }
    else:
        summary = None
    
    return {
        "import_id": import_id,
        "results": results_list,
        "summary": summary
    }


@app.get("/imports/{import_id}/notices/{notice_id}/sroi-results")
def get_notice_sroi_results(
    import_id: str, 
    notice_id: str, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Haal SROI resultaat op voor een specifieke notice.
    """
    # Verify import ownership
    import_record = db.query(Import)\
        .filter(Import.id == import_id)\
        .first()
    
    if not import_record:
        raise HTTPException(status_code=404, detail="Import niet gevonden")
    if import_record.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toegang tot deze import")
    
    # Fetch result for this notice
    result = db.query(NoticeSROIResult)\
        .filter(NoticeSROIResult.notice_id == notice_id)\
        .filter(NoticeSROIResult.owner_code == user_code)\
        .first()
    
    if not result:
        return {
            "notice_id": notice_id,
            "result": None,
            "message": "Nog geen analyse beschikbaar voor deze notice"
        }
    
    # Convert to dict
    result_dict = result.__dict__
    result_dict.pop('_sa_instance_state', None)
    
    return {
        "notice_id": notice_id,
        "result": result_dict
    }


@app.delete("/imports/{import_id}/sroi-results")
def delete_sroi_results(
    import_id: str, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Verwijder SROI resultaten (om opnieuw te analyseren).
    """
    # Ensure import belongs to user
    import_record = db.query(Import)\
        .filter(Import.id == import_id)\
        .first()
    
    if not import_record:
        raise HTTPException(status_code=404, detail="Import niet gevonden")
    if import_record.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toegang tot deze import")

    # Delete results
    deleted_count = db.query(SROIResult)\
        .filter(SROIResult.import_id == import_id)\
        .filter(SROIResult.owner_code == user_code)\
        .delete()
    
    db.commit()
    
    # Clear status
    if import_id in sroi_analysis_status:
        del sroi_analysis_status[import_id]
    
    return {
        "message": f"{deleted_count} SROI resultaten verwijderd",
        "import_id": import_id
    }


@app.get("/imports/{import_id}/notices/{notice_id}")
def get_notice_detail(
    import_id: str, 
    notice_id: str, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Fetch a single notice detail by import_id and notice_id.
    """
    print(f"\n=== FETCH NOTICE DETAIL ===")
    print(f"Import ID: {import_id!r}")
    print(f"Notice ID: {notice_id!r}")
    
    try:
        # Query notices table for the specific record
        notice = db.query(Notice)\
            .filter(Notice.import_id == import_id)\
            .filter(Notice.id == notice_id)\
            .filter(Notice.owner_code == user_code)\
            .first()
        
        if not notice:
            print(f"❌ Notice not found: import_id={import_id!r}, notice_id={notice_id!r}")
            raise HTTPException(
                status_code=404, 
                detail=f"Notice not found for import_id={import_id} and notice_id={notice_id}"
            )
        
        # Convert to dict
        notice_dict = notice.__dict__
        notice_dict.pop('_sa_instance_state', None)
        
        print(f"✅ Found notice: {notice_dict.get('titel', 'No title')}")
        return notice_dict
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR fetching notice detail: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch notice detail: {str(e)}"
        )


@app.post("/imports/{import_id}/notices/{notice_id}/sroi-analyze")
def analyze_single_notice(
    import_id: str,
    notice_id: str,
    payload: Optional[SingleNoticeSROIRequest] = None,
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Analyseer één specifieke notice / bedrijf (winnaar of inkoper) en sla resultaat direct op.
    Body (optioneel): { "target": "winner" | "buyer" }

    NOTE: this is deliberately `def`, not `async def`. It calls analyze_notice_sroi,
    which does blocking HTTP (up to 7 page fetches) plus time.sleep() and can run for
    60-90s. As an async handler it blocked the event loop for that whole time, stalling
    every other request in the process -- health checks included. As a sync handler
    Starlette runs it in the threadpool instead. The real fix is a job queue (Phase 1).
    """
    try:
        if not notice_id:
            raise HTTPException(status_code=400, detail="Notice ID is missing or invalid")

        target = payload.target if payload else "winner"

        # Verify import exists and ownership
        import_record = db.query(Import)\
            .filter(Import.id == import_id)\
            .first()
        
        if not import_record:
            raise HTTPException(status_code=404, detail="Import niet gevonden")
        if import_record.owner_code != user_code:
            raise HTTPException(status_code=403, detail="Je hebt geen toegang tot deze import")

        # Fetch the notice record
        notice = db.query(Notice)\
            .filter(Notice.import_id == import_id)\
            .filter(Notice.id == notice_id)\
            .filter(Notice.owner_code == user_code)\
            .first()

        if not notice:
            raise HTTPException(status_code=404, detail="Notice niet gevonden")

        # Convert to dict for analysis
        notice_dict = notice.__dict__
        notice_dict.pop('_sa_instance_state', None)

        # If caller requested the buyer specifically, coerce fields so analyze_notice_sroi targets buyer
        notice_for_analysis = dict(notice_dict)
        if target == "buyer":
            notice_for_analysis["win_bedrijf_naam"] = notice_dict.get("buyer_bedrijf_naam")
            notice_for_analysis["win_website"] = notice_dict.get("buyer_website")

        # Run analysis
        try:
            result = analyze_notice_sroi(notice_for_analysis)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Analyse mislukt: {e}")

        # Attach metadata
        result["notice_id"] = notice_id
        result["target"] = target
        
        # Save using new function
        save_notice_sroi_result(result, owner_code=user_code, db=db)

        return result

    except HTTPException:
        raise
    except Exception as e:
        print("ERROR in analyze_single_notice:", e)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/imports/{import_id}/sroi-download")
def download_sroi_results(
    import_id: str,
    format: str = Query("excel", pattern="^(excel|csv)$"),
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """
    Download SROI resultaten als Excel of CSV.
    """
    # Scope to the caller: this endpoint previously had no auth and no owner
    # filter, so any caller who knew an import_id could download its results.
    results = db.query(SROIResult)\
        .filter(SROIResult.import_id == import_id)\
        .filter(SROIResult.owner_code == user_code)\
        .order_by(desc(SROIResult.score))\
        .all()
    
    if not results:
        raise HTTPException(404, "Geen SROI resultaten voor deze import")
    
    # Convert to dict
    rows = []
    for result in results:
        row = result.__dict__.copy()
        row.pop('_sa_instance_state', None)
        
        # Flatten evidence array voor export
        if isinstance(row.get("evidence"), list):
            row["evidence"] = ", ".join(row["evidence"])
        
        rows.append(row)
    
    if format == "csv":
        return _make_csv_response(rows, filename=f"sroi_{import_id}.csv")
    else:
        return _make_excel_response(rows, filename=f"sroi_{import_id}.xlsx")


# ========================================
# REGIONS ENDPOINT
# ========================================

@app.get("/regions")
def list_regions(include_cities: bool = Query(False, description="Include city lists per province")):
    """
    Return available provinces (regio's). Set include_cities=true to also return example cities per province.
    """
    if include_cities:
        return {"regions": PROVINCE_TO_CITIES}
    return {"regions": list(PROVINCE_TO_CITIES.keys())}


@app.get("/analytics/buyers")
def buyer_spend_profile(limit: int = Query(50, ge=1, le=200), db: Session = Depends(get_db), user_code: str = Depends(validate_user_code)):
    """Buyer profiles from canonical public procurement data (not owner data)."""
    rows = db.query(
        TenderCompany.name, TenderCompany.kvk, TenderCompany.province,
        func.count(TenderNotice.id).label("awards"),
        func.max(TenderNotice.publicatie_datum).label("latest_award"),
    ).join(TenderNotice, TenderNotice.buyer_company_id == TenderCompany.id).group_by(
        TenderCompany.id
    ).order_by(func.count(TenderNotice.id).desc()).limit(limit).all()
    return [{"buyer": row.name, "kvk": row.kvk, "province": row.province, "awards": row.awards, "latest_award": row.latest_award} for row in rows]


@app.get("/saved-searches")
def list_saved_searches(user_code: str = Depends(validate_user_code), db: Session = Depends(get_db)):
    return [{"id": row.id, "name": row.name, "filters": row.filters, "alert_enabled": row.is_alert_enabled,
             "last_checked_at": row.last_checked_at, "created_at": row.created_at}
            for row in db.query(SavedSearch).filter(SavedSearch.owner_code == user_code).order_by(SavedSearch.created_at.desc()).all()]


@app.post("/saved-searches", status_code=201)
def create_saved_search(payload: dict, user_code: str = Depends(validate_user_code), db: Session = Depends(get_db)):
    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name is required")
    row = SavedSearch(owner_code=user_code, name=name, filters=payload.get("filters") or {}, is_alert_enabled=payload.get("alert_enabled", True))
    db.add(row)
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(409, "A saved search with this name already exists")
    db.refresh(row)
    return {"id": row.id, "name": row.name, "filters": row.filters, "alert_enabled": row.is_alert_enabled}


@app.delete("/saved-searches/{search_id}", status_code=204)
def delete_saved_search(search_id: int, user_code: str = Depends(validate_user_code), db: Session = Depends(get_db)):
    row = db.query(SavedSearch).filter(SavedSearch.id == search_id, SavedSearch.owner_code == user_code).first()
    if not row:
        raise HTTPException(404, "Saved search not found")
    db.delete(row)
    db.commit()


@app.get("/analytics/competitors")
def competitor_wins(limit: int = Query(50, ge=1, le=200), db: Session = Depends(get_db), user_code: str = Depends(validate_user_code)):
    """Rank suppliers by canonical award count; filters cancelled/VEAT records."""
    rows = db.query(
        TenderCompany.name, TenderCompany.kvk, TenderCompany.province,
        func.count(TenderNotice.id).label("wins"), func.max(TenderNotice.publicatie_datum).label("latest_win"),
    ).join(TenderNotice, TenderNotice.winner_company_id == TenderCompany.id).filter(
        TenderNotice.record_type == "awards", TenderNotice.is_cancelled.is_(False)
    ).group_by(TenderCompany.id).order_by(func.count(TenderNotice.id).desc()).limit(limit).all()
    return [{"company": row.name, "kvk": row.kvk, "province": row.province, "wins": row.wins, "latest_win": row.latest_win} for row in rows]


@app.get("/analytics/re-tenders")
def re_tender_candidates(within_days: int = Query(180, ge=1, le=3650), db: Session = Depends(get_db), user_code: str = Depends(validate_user_code)):
    """Lots whose known contract end date is approaching; excludes unknown durations."""
    today = dt.date.today()
    cutoff = today + dt.timedelta(days=within_days)
    lots = db.query(TenderNoticeLot, TenderNotice, TenderCompany).join(
        TenderNotice, TenderNoticeLot.notice_id == TenderNotice.id
    ).outerjoin(TenderCompany, TenderNoticeLot.awarded_company_id == TenderCompany.id).filter(
        TenderNoticeLot.contract_end >= today, TenderNoticeLot.contract_end <= cutoff
    ).order_by(TenderNoticeLot.contract_end).all()
    return [{"notice_id": notice.notice_id, "lot_id": lot.external_lot_id, "title": lot.title or notice.title,
             "contract_end": lot.contract_end, "winner": company.name if company else None,
             "days_until_end": (lot.contract_end - today).days} for lot, notice, company in lots]


# ========================================
# CRM ENDPOINTS
# ========================================

CRM_COMPANIES_CREATE_SQL = """
CREATE TABLE IF NOT EXISTS crm_companies (
  id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
  name TEXT NOT NULL,
  owner_code TEXT,
  website TEXT,
  kvk TEXT,
  contact_name TEXT,
  contact_email TEXT,
  contact_phone TEXT,
  source_notice_id TEXT,
  lead_status TEXT DEFAULT 'new', -- e.g. new, contacted, interested, not_interested
  last_contacted TIMESTAMPTZ,
  notes TEXT,
  extra JSONB,
  created_at TIMESTAMPTZ DEFAULT NOW(),
  updated_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_crm_companies_name ON crm_companies (lower(name));
CREATE INDEX IF NOT EXISTS idx_crm_companies_owner ON crm_companies (owner_code);
"""

CRM_FOLLOWUPS_CREATE_SQL = """
CREATE TABLE IF NOT EXISTS crm_followups (
  id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
  company_id BIGINT REFERENCES crm_companies(id) ON DELETE CASCADE,
  scheduled_at TIMESTAMPTZ,
  completed BOOLEAN DEFAULT FALSE,
  emailed BOOLEAN DEFAULT FALSE,
  action TEXT,
  note TEXT,
  created_by TEXT,
  created_at TIMESTAMPTZ DEFAULT NOW(),
  updated_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_crm_followups_company ON crm_followups (company_id);
"""


@app.get("/crm/init")
def crm_init():
    """Return SQL to create CRM tables. Run these statements in Supabase SQL editor if tables don't exist."""
    return {
        "message": "If your database doesn't have the CRM tables, run the provided SQL in your SQL editor or use Alembic migrations.",
        "crm_companies_sql": CRM_COMPANIES_CREATE_SQL,
        "crm_followups_sql": CRM_FOLLOWUPS_CREATE_SQL,
    }


def _ensure_crm_tables_ok(db: Session):
    """Quick check whether crm_companies exists; return (ok, error_message)"""
    try:
        # Try a simple query to check if table exists
        db.query(CRMCompany).limit(1).all()
        return True, None
    except Exception as e:
        return False, str(e)


@app.get("/crm/companies")
def crm_list_companies(
    q: Optional[str] = Query(None, description="Search by name substring"), 
    status: Optional[str] = Query(None), 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """List companies (leads). Optional search by name and filter by lead_status."""
    ok, err = _ensure_crm_tables_ok(db)
    if not ok:
        raise HTTPException(
            status_code=500, 
            detail={
                "error": "CRM tables not found", 
                "advice": "Call /crm/init and run SQL or create migrations", 
                "raw": err
            }
        )

    query = db.query(CRMCompany).filter(CRMCompany.owner_code == user_code)
    
    if q:
        # Case-insensitive search
        query = query.filter(func.lower(CRMCompany.name).like(f"%{q.lower()}%"))
    
    if status:
        query = query.filter(CRMCompany.lead_status == status)

    companies = query.order_by(desc(CRMCompany.updated_at)).limit(100).all()
    
    # Convert to dict
    result = []
    for company in companies:
        company_dict = company.__dict__.copy()
        company_dict.pop('_sa_instance_state', None)
        result.append(company_dict)
    
    return result


@app.post("/crm/companies")
def crm_create_company(
    payload: dict, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """Create a new company lead. Expected JSON payload keys: name (required), website, kvk, contact_name, contact_email, source_notice_id, notes, lead_status"""
    ok, err = _ensure_crm_tables_ok(db)
    if not ok:
        raise HTTPException(
            status_code=500, 
            detail={
                "error": "CRM tables not found", 
                "advice": "Call /crm/init and run SQL or create migrations", 
                "raw": err
            }
        )

    if not payload.get("name"):
        raise HTTPException(status_code=400, detail="Field 'name' is required")

    # Create new company
    new_company = CRMCompany(
        name=payload.get("name"),
        owner_code=user_code,
        website=payload.get("website"),
        kvk=payload.get("kvk"),
        contact_name=payload.get("contact_name"),
        contact_email=payload.get("contact_email"),
        contact_phone=payload.get("contact_phone"),
        source_notice_id=payload.get("source_notice_id"),
        lead_status=payload.get("lead_status", "new"),
        last_contacted=payload.get("last_contacted"),
        notes=payload.get("notes"),
        extra=payload.get("extra"),
        created_at=dt.datetime.utcnow(),
        updated_at=dt.datetime.utcnow()
    )
    
    db.add(new_company)
    db.commit()
    db.refresh(new_company)
    
    # Convert to dict
    result = new_company.__dict__.copy()
    result.pop('_sa_instance_state', None)
    
    return result


@app.get("/crm/companies/{company_id}")
def crm_get_company(
    company_id: int, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    ok, err = _ensure_crm_tables_ok(db)
    if not ok:
        raise HTTPException(
            status_code=500, 
            detail={
                "error": "CRM tables not found", 
                "advice": "Call /crm/init and run SQL or create migrations", 
                "raw": err
            }
        )

    company = db.query(CRMCompany).filter(CRMCompany.id == company_id).first()
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    if company.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toegang tot dit bedrijf")

    # Fetch followups
    followups = db.query(CRMFollowup)\
        .filter(CRMFollowup.company_id == company_id)\
        .order_by(CRMFollowup.scheduled_at)\
        .all()
    
    # Convert to dict
    company_dict = company.__dict__.copy()
    company_dict.pop('_sa_instance_state', None)
    
    followups_list = []
    for followup in followups:
        followup_dict = followup.__dict__.copy()
        followup_dict.pop('_sa_instance_state', None)
        followups_list.append(followup_dict)
    
    company_dict["followups"] = followups_list
    
    return company_dict


@app.patch("/crm/companies/{company_id}")
def crm_update_company(
    company_id: int, 
    payload: dict, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    ok, err = _ensure_crm_tables_ok(db)
    if not ok:
        raise HTTPException(
            status_code=500, 
            detail={
                "error": "CRM tables not found", 
                "advice": "Call /crm/init and run SQL or create migrations", 
                "raw": err
            }
        )

    # Ensure company belongs to user
    company = db.query(CRMCompany).filter(CRMCompany.id == company_id).first()
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    if company.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toestemming om dit bedrijf te bewerken")

    # Update fields
    for key, value in payload.items():
        if hasattr(company, key) and key not in ['id', 'created_at', 'owner_code']:
            setattr(company, key, value)
    
    company.updated_at = dt.datetime.utcnow()
    
    db.commit()
    db.refresh(company)
    
    # Convert to dict
    result = company.__dict__.copy()
    result.pop('_sa_instance_state', None)
    
    return result


@app.post("/crm/companies/{company_id}/followups")
def crm_add_followup(
    company_id: int, 
    payload: dict, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    """Add a follow-up action/note for a company. payload keys: scheduled_at (ISO), action, note, emailed (bool), completed (bool), created_by"""
    ok, err = _ensure_crm_tables_ok(db)
    if not ok:
        raise HTTPException(
            status_code=500, 
            detail={
                "error": "CRM tables not found", 
                "advice": "Call /crm/init and run SQL or create migrations", 
                "raw": err
            }
        )

    # Validate company exists and belongs to user
    company = db.query(CRMCompany)\
        .filter(CRMCompany.id == company_id)\
        .first()
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    if company.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toestemming om followups voor dit bedrijf toe te voegen")

    # Create new followup
    new_followup = CRMFollowup(
        company_id=company_id,
        scheduled_at=payload.get("scheduled_at"),
        action=payload.get("action"),
        note=payload.get("note"),
        emailed=payload.get("emailed", False),
        completed=payload.get("completed", False),
        created_by=payload.get("created_by"),
        created_at=dt.datetime.utcnow(),
        updated_at=dt.datetime.utcnow()
    )
    
    db.add(new_followup)
    db.commit()
    db.refresh(new_followup)
    
    # Convert to dict
    result = new_followup.__dict__.copy()
    result.pop('_sa_instance_state', None)
    
    return result


@app.get("/crm/companies/{company_id}/followups")
def crm_list_followups(
    company_id: int, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    ok, err = _ensure_crm_tables_ok(db)
    if not ok:
        raise HTTPException(
            status_code=500, 
            detail={
                "error": "CRM tables not found", 
                "advice": "Call /crm/init and run SQL or create migrations", 
                "raw": err
            }
        )

    # Ensure company belongs to user
    company = db.query(CRMCompany)\
        .filter(CRMCompany.id == company_id)\
        .first()
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    if company.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toegang tot de followups van dit bedrijf")

    followups = db.query(CRMFollowup)\
        .filter(CRMFollowup.company_id == company_id)\
        .order_by(CRMFollowup.scheduled_at)\
        .all()
    
    # Convert to dict
    result = []
    for followup in followups:
        followup_dict = followup.__dict__.copy()
        followup_dict.pop('_sa_instance_state', None)
        result.append(followup_dict)
    
    return result


@app.patch("/crm/followups/{followup_id}")
def crm_update_followup(
    followup_id: int, 
    payload: dict, 
    user_code: str = Depends(validate_user_code),
    db: Session = Depends(get_db)
):
    ok, err = _ensure_crm_tables_ok(db)
    if not ok:
        raise HTTPException(
            status_code=500, 
            detail={
                "error": "CRM tables not found", 
                "advice": "Call /crm/init and run SQL or create migrations", 
                "raw": err
            }
        )

    # Ensure followup belongs to a company owned by user
    followup = db.query(CRMFollowup)\
        .filter(CRMFollowup.id == followup_id)\
        .first()
    
    if not followup:
        raise HTTPException(status_code=404, detail="Followup not found")
    
    company = db.query(CRMCompany)\
        .filter(CRMCompany.id == followup.company_id)\
        .first()
    
    if not company or company.owner_code != user_code:
        raise HTTPException(status_code=403, detail="Je hebt geen toestemming om deze followup te bewerken")

    # Update fields
    for key, value in payload.items():
        if hasattr(followup, key) and key not in ['id', 'company_id', 'created_at']:
            setattr(followup, key, value)
    
    followup.updated_at = dt.datetime.utcnow()
    
    db.commit()
    db.refresh(followup)
    
    # Convert to dict
    result = followup.__dict__.copy()
    result.pop('_sa_instance_state', None)
    
    return result


# Runnen:
# uvicorn serve:app --reload --port 8000
